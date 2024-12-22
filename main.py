from openpyxl import load_workbook
import statistics as st
import matplotlib as fig


workbook = load_workbook(filename='EstDad.xlsx')


def get_data(sheet_name):
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

